# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('/Users/danisouza/Documents/Registration.xlsx')

# create the sheet object
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "First Name"
    sheet.cell(row=1, column=2).value = "Last Name"
    sheet.cell(row=1, column=3).value = "Contact Number"
    sheet.cell(row=1, column=4).value = "Email id"
    sheet.cell(row=1, column=5).value = "Neighborhood"


# Function to set focus (cursor)
def focus1():
    # set focus on the firstname_field box
    first_name.focus_set()


# Function to set focus
def focus2():
    # set focus on the last_name box
    last_name.focus_set()


# Function to set focus
def focus3():
    # set focus on the contact_number box
    contact_number.focus_set()


# Function to set focus
def focus4():
    # set focus on the email_id box
    email_id.focus_set()


# Function to set focus
def focus5():
    # set focus on the neighborhood box
    neighborhood.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    first_name.delete(0, END)
    last_name.delete(0, END)
    contact_number.delete(0, END)
    email_id.delete(0, END)
    neighborhood.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (first_name.get() == "" and
            last_name.get() == "" and
            contact_number.get() == "" and
            email_id.get() == "" and
            neighborhood.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = first_name.get()
        sheet.cell(row=current_row + 1, column=2).value = last_name.get()
        sheet.cell(row=current_row + 1, column=3).value = contact_number.get()
        sheet.cell(row=current_row + 1, column=4).value = email_id.get()
        sheet.cell(row=current_row + 1, column=7).value = neighborhood.get()

        # save the file
        wb.save('/Users/danisouza/Documents/Registration.xlsx')

        # set focus on the name_field box
        first_name.focus_set()

        # call the clear() function
        clear()

    # Driver code


if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("500x300")

    excel()

    # create a Form label
    heading = Label(root, text="Form", bg="light green")

    # create a First Name label
    first_name = Label(root, text="First Name", bg="light green")

    # create a Last Name label
    last_name = Label(root, text="Last Name", bg="light green")

    # create a Contact No. label
    contact_number = Label(root, text="Contact No.", bg="light green")

    # create a Email id label
    email_id = Label(root, text="Email id", bg="light green")

    # create a Neighborhood label
    neighborhood = Label(root, text="Neighborhood", bg="light green")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    first_name.grid(row=1, column=0)
    last_name.grid(row=2, column=0)
    contact_number.grid(row=3, column=0)
    email_id.grid(row=4, column=0)
    neighborhood.grid(row=5, column=0)

    # create a text entry box
    # for typing the information
    first_name = Entry(root)
    last_name = Entry(root)
    contact_number = Entry(root)
    email_id = Entry(root)
    neighborhood = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    first_name.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    last_name.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    contact_number.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    email_id.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    neighborhood.bind("<Return>", focus5)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    first_name.grid(row=1, column=1, ipadx="100")
    last_name.grid(row=2, column=1, ipadx="100")
    contact_number.grid(row=3, column=1, ipadx="100")
    email_id.grid(row=4, column=1, ipadx="100")
    neighborhood.grid(row=5, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=6, column=1)

    # start the GUI
    root.mainloop()
